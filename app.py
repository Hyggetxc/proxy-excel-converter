from __future__ import annotations

import io
import json
import os
import socketserver
import traceback
import time
from cgi import FieldStorage
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler
from ipaddress import ip_address
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
REQUIRED_HEADERS = ["IP地址", "代理地址:端口号", "账号", "密码"]
EXPORT_HEADERS = ["IP地址", "代理地址", "代理端口", "代理账号", "代理密码", "备注"]
CONNECTIVITY_TIMEOUT_SECONDS = 8
CONNECTIVITY_MAX_ROWS = 100
CONNECTIVITY_TARGETS = [
    {
        "key": "google",
        "name": "Google",
        "url": "https://www.gstatic.com/generate_204",
        "expected_status": 204,
        "expected_text": None,
    },
    {
        "key": "cloudflare",
        "name": "Cloudflare",
        "url": "https://cp.cloudflare.com/generate_204",
        "expected_status": 204,
        "expected_text": None,
    },
    {
        "key": "apple",
        "name": "Apple",
        "url": "https://captive.apple.com/hotspot-detect.html",
        "expected_status": 200,
        "expected_text": "Success",
    },
]


@dataclass
class ValidationErrorItem:
    row: int
    message: str


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def normalize_cell(value: Any) -> str:
    if value is None:
      return ""
    return str(value).strip()


def validate_ip(value: str) -> bool:
    try:
        ip_address(value)
        return True
    except ValueError:
        return False


def split_host_port(value: str) -> tuple[str, str] | None:
    text = value.strip()
    if ":" not in text:
        return None
    host, port = text.rsplit(":", 1)
    host = host.strip()
    port = port.strip()
    if not host or not port:
        return None
    return host, port


def parse_workbook(file_bytes: bytes) -> dict[str, Any]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.worksheets[0]

    header_values = [normalize_cell(cell.value) for cell in sheet[1]]
    header_map = {value: index for index, value in enumerate(header_values) if value}
    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]

    if missing_headers:
        return {
            "ok": False,
            "sheet_name": sheet.title,
            "summary": {
                "total": 0,
                "valid": 0,
                "errors": len(missing_headers),
                "skipped_empty": 0,
            },
            "rows": [],
            "errors": [
                {
                    "row": 1,
                    "message": f"缺少表头：{header}",
                }
                for header in missing_headers
            ],
        }

    rows: list[dict[str, Any]] = []
    errors: list[ValidationErrorItem] = []
    valid_count = 0
    non_empty_count = 0
    skipped_empty_count = 0

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [normalize_cell(cell) for cell in row]
        row_has_any_value = any(values)
        if not row_has_any_value:
            skipped_empty_count += 1
            continue

        non_empty_count += 1

        ip_value = values[header_map["IP地址"]] if header_map["IP地址"] < len(values) else ""
        proxy_value = values[header_map["代理地址:端口号"]] if header_map["代理地址:端口号"] < len(values) else ""
        username = values[header_map["账号"]] if header_map["账号"] < len(values) else ""
        password = values[header_map["密码"]] if header_map["密码"] < len(values) else ""
        remark = values[header_map["备注"]] if "备注" in header_map and header_map["备注"] < len(values) else ""

        row_errors: list[str] = []

        if not ip_value:
            row_errors.append("缺少必填字段 `IP地址`。")
        elif not validate_ip(ip_value):
            row_errors.append("`IP地址` 不是有效的 IP。")

        if not proxy_value:
            row_errors.append("缺少必填字段 `代理地址:端口号`。")
            host = ""
            port_text = ""
        else:
            split_value = split_host_port(proxy_value)
            if not split_value:
                row_errors.append("`代理地址:端口号` 无法拆分成地址与端口。")
                host = ""
                port_text = ""
            else:
                host, port_text = split_value
                if not port_text.isdigit():
                    row_errors.append("`代理端口` 不是有效数字。")
                else:
                    port_value = int(port_text)
                    if port_value < 1 or port_value > 65535:
                        row_errors.append(f"`代理端口` 超出有效范围，当前值为 `{port_text}`。")

        if not username:
            row_errors.append("缺少必填字段 `账号`。")

        if not password:
            row_errors.append("缺少必填字段 `密码`。")

        status = "通过" if not row_errors else "错误"
        if status == "通过":
            valid_count += 1
        else:
            for message in row_errors:
                errors.append(ValidationErrorItem(row=row_number, message=message))

        rows.append(
            {
                "ip": ip_value,
                "host": host,
                "port": port_text,
                "username": username,
                "password": password,
                "remark": remark,
                "status": status,
            }
        )

    return {
        "ok": len(errors) == 0,
        "sheet_name": sheet.title,
        "summary": {
            "total": non_empty_count,
            "valid": valid_count,
            "errors": len(errors),
            "skipped_empty": skipped_empty_count,
        },
        "rows": rows,
        "errors": [{"row": item.row, "message": item.message} for item in errors],
    }


def export_workbook(validation_result: dict[str, Any]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(EXPORT_HEADERS)

    for row in validation_result["rows"]:
        sheet.append(
            [
                row["ip"],
                row["host"],
                int(row["port"]),
                row["username"],
                row["password"],
                row["remark"],
            ]
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def build_socks_proxy_url(row: dict[str, Any]) -> str:
    host = normalize_cell(row.get("host"))
    port = normalize_cell(row.get("port"))
    username = quote(normalize_cell(row.get("username")), safe="")
    password = quote(normalize_cell(row.get("password")), safe="")
    return f"socks5h://{username}:{password}@{host}:{port}"


def check_connectivity_target(row: dict[str, Any], target: dict[str, Any]) -> dict[str, Any]:
    try:
        import requests
    except ImportError:
        return {"key": target["key"], "ok": False, "label": "缺少依赖", "error": "missing_requests"}

    proxy_url = build_socks_proxy_url(row)
    started_at = time.perf_counter()

    try:
        response = requests.get(
            target["url"],
            proxies={"http": proxy_url, "https": proxy_url},
            timeout=CONNECTIVITY_TIMEOUT_SECONDS,
            allow_redirects=False,
            headers={"User-Agent": "proxy-excel-converter/1.0"},
        )
        latency_ms = int((time.perf_counter() - started_at) * 1000)
        expected_status = target["expected_status"]
        expected_text = target["expected_text"]
        body = response.text[:256] if expected_text else ""
        ok = response.status_code == expected_status and (not expected_text or expected_text in body)

        if ok:
            label = f"可用 {latency_ms}ms"
        elif response.status_code in {401, 407}:
            label = "认证失败"
        else:
            label = f"状态异常 {response.status_code}"

        return {
            "key": target["key"],
            "ok": ok,
            "label": label,
            "latencyMs": latency_ms,
            "status": response.status_code,
        }
    except requests.exceptions.Timeout:
        return {"key": target["key"], "ok": False, "label": "超时", "error": "timeout"}
    except requests.exceptions.ProxyError:
        return {"key": target["key"], "ok": False, "label": "代理失败", "error": "proxy_error"}
    except requests.exceptions.ConnectionError:
        return {"key": target["key"], "ok": False, "label": "连接失败", "error": "connection_error"}
    except requests.exceptions.RequestException:
        return {"key": target["key"], "ok": False, "label": "失败", "error": "request_error"}


def validate_connectivity_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    valid_rows: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=1):
        row_number = row.get("row") or index
        host = normalize_cell(row.get("host"))
        port = normalize_cell(row.get("port"))
        username = normalize_cell(row.get("username"))
        password = normalize_cell(row.get("password"))

        if not host or not port or not username or not password:
            raise ValueError(f"第 {row_number} 行缺少代理地址、端口、账号或密码。")

        if not port.isdigit() or not 1 <= int(port) <= 65535:
            raise ValueError(f"第 {row_number} 行代理端口无效。")

        valid_rows.append(
            {
                "row": int(row_number),
                "host": host,
                "port": port,
                "username": username,
                "password": password,
            }
        )

    return valid_rows


def check_connectivity(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if len(rows) > CONNECTIVITY_MAX_ROWS:
        raise ValueError(f"单次最多检测 {CONNECTIVITY_MAX_ROWS} 条代理。")

    valid_rows = validate_connectivity_rows(rows)
    result_map: dict[int, dict[str, Any]] = {
        row["row"]: {
            "row": row["row"],
            "targets": {
                target["key"]: {"ok": False, "label": "待检测"}
                for target in CONNECTIVITY_TARGETS
            },
        }
        for row in valid_rows
    }

    with ThreadPoolExecutor(max_workers=min(12, max(1, len(valid_rows) * len(CONNECTIVITY_TARGETS)))) as executor:
        future_map = {}
        for row in valid_rows:
            for target in CONNECTIVITY_TARGETS:
                future = executor.submit(check_connectivity_target, row, target)
                future_map[future] = row["row"]

        for future in as_completed(future_map):
            row_number = future_map[future]
            target_result = future.result()
            result_map[row_number]["targets"][target_result["key"]] = target_result

    return {
        "ok": True,
        "targets": [{"key": target["key"], "name": target["name"]} for target in CONNECTIVITY_TARGETS],
        "results": [result_map[key] for key in sorted(result_map)],
    }


class ProxyExcelHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def log_message(self, format: str, *args: Any) -> None:
        return

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self._send_json({"ok": True})
            return
        return super().do_GET()

    def do_OPTIONS(self) -> None:
        self.send_response(HTTPStatus.NO_CONTENT)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/check-connectivity":
            self._handle_connectivity_check()
            return

        if parsed.path not in {"/api/validate", "/api/export"}:
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        try:
            form = self._parse_form_data()
            upload = form["file"] if "file" in form else None
            if upload is None or not getattr(upload, "file", None):
                self._send_json({"ok": False, "message": "没有收到上传文件。"}, status=HTTPStatus.BAD_REQUEST)
                return

            file_bytes = upload.file.read()
            validation = parse_workbook(file_bytes)

            if parsed.path == "/api/validate":
                payload = {
                    "ok": validation["ok"],
                    "file_name": upload.filename or "uploaded.xlsx",
                    "sheet_name": validation["sheet_name"],
                    "summary": validation["summary"],
                    "rows": validation["rows"],
                    "errors": validation["errors"],
                }
                self._send_json(payload, status=HTTPStatus.OK if validation["ok"] else HTTPStatus.UNPROCESSABLE_ENTITY)
                return

            if not validation["ok"]:
                self._send_json(
                    {
                        "ok": False,
                        "message": "校验未通过，无法导出。",
                        "sheet_name": validation["sheet_name"],
                        "summary": validation["summary"],
                        "errors": validation["errors"],
                    },
                    status=HTTPStatus.UNPROCESSABLE_ENTITY,
                )
                return

            output = export_workbook(validation)
            filename = "proxysheet-export.xlsx"
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Length", str(len(output)))
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.end_headers()
            self.wfile.write(output)
        except Exception as exc:  # noqa: BLE001
            self._send_json(
                {
                    "ok": False,
                    "message": f"服务处理失败：{exc}",
                    "trace": traceback.format_exc(),
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    def _handle_connectivity_check(self) -> None:
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
            rows = payload.get("rows")

            if not isinstance(rows, list) or not rows:
                self._send_json({"ok": False, "message": "没有收到可检测的代理数据。"}, status=HTTPStatus.BAD_REQUEST)
                return

            self._send_json(check_connectivity(rows))
        except ValueError as exc:
            self._send_json({"ok": False, "message": str(exc)}, status=HTTPStatus.BAD_REQUEST)
        except Exception as exc:  # noqa: BLE001
            self._send_json(
                {
                    "ok": False,
                    "message": f"连通性检测失败：{exc}",
                },
                status=HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    def _parse_form_data(self) -> FieldStorage:
        environ = {
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
        }
        return FieldStorage(fp=self.rfile, headers=self.headers, environ=environ)

    def _send_json(self, payload: dict[str, Any], status: int = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8765"))
    with socketserver.ThreadingTCPServer((host, port), ProxyExcelHandler) as httpd:
        print(f"Proxy Excel Converter running at http://{host}:{port}")
        httpd.serve_forever()


if __name__ == "__main__":
    main()
